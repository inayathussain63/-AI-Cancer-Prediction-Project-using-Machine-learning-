import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from sklearn.preprocessing import LabelEncoder, StandardScaler
from sklearn.model_selection import train_test_split
from sklearn.ensemble import RandomForestClassifier, GradientBoostingClassifier
from sklearn.linear_model import LogisticRegression
from sklearn.svm import SVC
from sklearn.neighbors import KNeighborsClassifier
from sklearn.metrics import accuracy_score, precision_score, recall_score, f1_score, confusion_matrix, roc_auc_score, roc_curve
import warnings
import os
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage

warnings.filterwarnings('ignore')

def generate_dummy_data(filename='final_cleaned_data.csv'):
    print(f"Generating dummy data: {filename}...")
    n_samples = 1000
    data = {
        'Risk_Score': np.random.uniform(0, 100, n_samples),
        'Gender': np.random.choice(['Male', 'Female'], n_samples),
        'Smoking': np.random.choice(['Yes', 'No'], n_samples),
        'Alcohol_Consumption': np.random.choice(['High', 'Moderate', 'Low'], n_samples),
        'Diabetes': np.random.choice(['Yes', 'No'], n_samples),
        'Hypertension': np.random.choice(['Yes', 'No'], n_samples),
        'Heart_Disease': np.random.choice(['Yes', 'No'], n_samples),
        'Insurance_Type': np.random.choice(['Basic', 'Premium'], n_samples),
        'Age': np.random.randint(18, 90, n_samples),
        'Height_cm': np.random.uniform(150, 200, n_samples),
        'Weight_kg': np.random.uniform(50, 120, n_samples),
        'BMI': np.random.uniform(18, 35, n_samples),
        'Systolic_BP': np.random.randint(90, 180, n_samples),
        'Diastolic_BP': np.random.randint(60, 120, n_samples),
        'Heart_Rate': np.random.randint(50, 100, n_samples),
        'Temperature_F': np.random.uniform(97, 100, n_samples),
        'Blood_Sugar': np.random.uniform(70, 200, n_samples),
        'Cholesterol': np.random.uniform(150, 300, n_samples),
        'Hemoglobin': np.random.uniform(10, 18, n_samples),
        'Exercise_Hours_Week': np.random.uniform(0, 10, n_samples),
        'Hospital_Visits_Year': np.random.randint(0, 5, n_samples)
    }
    df = pd.DataFrame(data)
    df.to_csv(filename, index=False)
    return df

def prepare_features_and_target(data):
    print("Preparing features and target...")
    processed_data = data.copy()
    risk_threshold = processed_data['Risk_Score'].median()
    processed_data['High_Risk'] = (processed_data['Risk_Score'] > risk_threshold).astype(int)

    label_encoders = {}
    categorical_columns = ['Gender', 'Smoking', 'Alcohol_Consumption', 'Diabetes', 'Hypertension', 'Heart_Disease', 'Insurance_Type']
    for col in categorical_columns:
        if col in processed_data.columns:
            le = LabelEncoder()
            processed_data[col + '_encoded'] = le.fit_transform(processed_data[col].astype(str))
            label_encoders[col] = le

    feature_columns = ['Age', 'Height_cm', 'Weight_kg', 'BMI', 'Systolic_BP', 'Diastolic_BP', 'Heart_Rate', 'Temperature_F', 'Blood_Sugar', 'Cholesterol', 'Hemoglobin', 'Exercise_Hours_Week', 'Hospital_Visits_Year', 'Gender_encoded', 'Smoking_encoded', 'Alcohol_Consumption_encoded', 'Diabetes_encoded', 'Hypertension_encoded', 'Heart_Disease_encoded', 'Insurance_Type_encoded']
    available_features = [col for col in feature_columns if col in processed_data.columns]
    X = processed_data[available_features]
    y = processed_data['High_Risk']
    return X, y, label_encoders, processed_data

def scale_and_split_data(X, y):
    print("Scaling and splitting data...")
    X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=42, stratify=y)
    scaler = StandardScaler()
    X_train_scaled = scaler.fit_transform(X_train)
    X_test_scaled = scaler.transform(X_test)
    X_train_scaled = pd.DataFrame(X_train_scaled, columns=X.columns)
    X_test_scaled = pd.DataFrame(X_test_scaled, columns=X.columns)
    return X_train_scaled, X_test_scaled, y_train, y_test, scaler

def train_and_evaluate_models(X_train, X_test, y_train, y_test):
    print("Training and evaluating multiple models...")
    models = {
        "RandomForest": RandomForestClassifier(n_estimators=100, random_state=42),
        "GradientBoosting": GradientBoostingClassifier(random_state=42),
        "LogisticRegression": LogisticRegression(random_state=42),
        "SVM": SVC(probability=True, random_state=42),
        "KNN": KNeighborsClassifier()
    }

    results = []
    model_probs = {} # Store probabilities for ROC
    best_model_name = None
    best_accuracy = 0.0
    best_cm = None

    for name, model in models.items():
        print(f"Training {name}...")
        model.fit(X_train, y_train)
        y_pred = model.predict(X_test)
        y_prob = model.predict_proba(X_test)[:, 1] if hasattr(model, "predict_proba") else y_pred
        model_probs[name] = y_prob

        acc = accuracy_score(y_test, y_pred)
        prec = precision_score(y_test, y_pred, average='weighted')
        rec = recall_score(y_test, y_pred, average='weighted')
        f1 = f1_score(y_test, y_pred, average='weighted')
        auc = roc_auc_score(y_test, y_prob)

        results.append({"Model": name, "Accuracy": acc, "Precision": prec, "Recall": rec, "F1_Score": f1, "AUC": auc})

        if acc > best_accuracy:
            best_accuracy = acc
            best_model_name = name
            best_cm = confusion_matrix(y_test, y_pred)

    results_df = pd.DataFrame(results)
    return results_df, best_cm, best_model_name, model_probs

def generate_plots(results_df, model_probs, y_test, best_cm, best_model_name):
    print("Generating refined visualizations...")
    
    # 1. Grouped Bar Chart (Performance Comparison)
    try:
        plt.figure(figsize=(12, 6))
        results_df.plot(x='Model', y=['Accuracy', 'Precision', 'Recall', 'F1_Score'], kind='bar', width=0.8, figsize=(12, 6))
        plt.title('Model Performance Comparison')
        plt.ylabel('Score')
        plt.xticks(rotation=45)
        plt.ylim(0, 1.1)
        plt.grid(axis='y', linestyle='--', alpha=0.7)
        plt.legend(loc='upper right')
        plt.tight_layout()
        plt.savefig('plot_performance_comparison.png')
        plt.close()
    except Exception as e: print(f"Error plotting performance: {e}")

    # 2. ROC Curves
    try:
        plt.figure(figsize=(10, 8))
        for name, probs in model_probs.items():
            fpr, tpr, _ = roc_curve(y_test, probs)
            auc_val = roc_auc_score(y_test, probs)
            plt.plot(fpr, tpr, label=f'{name} (AUC = {auc_val:.3f})')
        
        plt.plot([0, 1], [0, 1], 'k--', label='Random')
        plt.xlabel('False Positive Rate')
        plt.ylabel('True Positive Rate')
        plt.title('ROC Curves')
        plt.legend()
        plt.grid(True, alpha=0.3)
        plt.tight_layout()
        plt.savefig('plot_roc_curves.png')
        plt.close()
    except Exception as e: print(f"Error plotting ROC: {e}")

    # 3. Accuracy Comparison Bar
    try:
        plt.figure(figsize=(10, 6))
        bars = plt.bar(results_df['Model'], results_df['Accuracy'], color=plt.cm.viridis(np.linspace(0, 1, len(results_df))))
        plt.title('Model Accuracy Comparison')
        plt.ylabel('Accuracy')
        plt.ylim(0, 1.1)
        plt.xticks(rotation=45)
        # Add labels
        for bar in bars:
            height = bar.get_height()
            plt.text(bar.get_x() + bar.get_width()/2., height,
                     f'{height:.3f}', ha='center', va='bottom')
        plt.tight_layout()
        plt.savefig('plot_accuracy_comparison.png')
        plt.close()
    except Exception as e: print(f"Error plotting Accuracy: {e}")

    # 4. AUC Comparison Bar
    try:
        plt.figure(figsize=(10, 6))
        bars = plt.bar(results_df['Model'], results_df['AUC'], color=plt.cm.plasma(np.linspace(0, 1, len(results_df))))
        plt.title('Model AUC Comparison')
        plt.ylabel('AUC Score')
        plt.ylim(0, 1.1)
        plt.xticks(rotation=45)
        for bar in bars:
            height = bar.get_height()
            plt.text(bar.get_x() + bar.get_width()/2., height,
                     f'{height:.3f}', ha='center', va='bottom')
        plt.tight_layout()
        plt.savefig('plot_auc_comparison.png')
        plt.close()
    except Exception as e: print(f"Error plotting AUC: {e}")

    # 5. Best Model Confusion Matrix
    try:
        plt.figure(figsize=(8, 6))
        plt.imshow(best_cm, interpolation='nearest', cmap=plt.cm.Blues)
        plt.title(f'Confusion Matrix - {best_model_name}')
        plt.colorbar()
        tick_marks = np.arange(2)
        plt.xticks(tick_marks, ['Predicted 0', 'Predicted 1'], rotation=45)
        plt.yticks(tick_marks, ['Actual 0', 'Actual 1'])
        
        thresh = best_cm.max() / 2.
        for i, j in np.ndindex(best_cm.shape):
            plt.text(j, i, format(best_cm[i, j], 'd'),
                     horizontalalignment="center",
                     color="white" if best_cm[i, j] > thresh else "black")
        
        plt.tight_layout()
        plt.savefig('plot_best_model_cm.png')
        plt.close()
    except Exception as e: print(f"Error plotting CM: {e}")

def save_artifacts(X_train, X_test, y_train, y_test, results_df, best_cm, best_model_name):
    print("Saving Excel artifacts...")
    train_df = pd.concat([X_train.reset_index(drop=True), y_train.reset_index(drop=True)], axis=1)
    train_df['Set'] = 'Train'
    test_df = pd.concat([X_test.reset_index(drop=True), y_test.reset_index(drop=True)], axis=1)
    test_df['Set'] = 'Test'
    full_processed = pd.concat([train_df, test_df], axis=0)
    full_processed.to_excel('AI_Makalu_ProcessedDataset.xlsx', index=False)

    cm_df = pd.DataFrame(best_cm, index=['Actual 0', 'Actual 1'], columns=['Predicted 0', 'Predicted 1'])
    with pd.ExcelWriter('AI_Makalu_ModelMetrics.xlsx', engine='openpyxl') as writer:
        results_df.to_excel(writer, sheet_name='Model Comparison', index=False)
        cm_df.to_excel(writer, sheet_name=f'Best Model ({best_model_name}) CM')
    
    # Embed images into Excel
    try:
        wb = load_workbook('AI_Makalu_ModelMetrics.xlsx')
        ws = wb.create_sheet('Visualizations')
        
        plot_files = [
            'plot_performance_comparison.png',
            'plot_roc_curves.png',
            'plot_accuracy_comparison.png',
            'plot_auc_comparison.png',
            'plot_best_model_cm.png'
        ]
        
        current_row = 1
        for plot_file in plot_files:
            if os.path.exists(plot_file):
                img = OpenpyxlImage(plot_file)
                cell_address = f'A{current_row}'
                ws.add_image(img, cell_address)
                # Estimate height (approx 20 pixels per row, image height ~400-600px -> ~20-30 rows)
                current_row += 35 
                
        wb.save('AI_Makalu_ModelMetrics.xlsx')
        print("Images embedded into AI_Makalu_ModelMetrics.xlsx")
    except Exception as e:
        print(f"Error embedding images to Excel: {e}")

    print("Artifacts saved: AI_Makalu_ProcessedDataset.xlsx, AI_Makalu_ModelMetrics.xlsx (with images), and plots.")

if __name__ == "__main__":
    try:
        if not os.path.exists('final_cleaned_data.csv'):
            print("Input file 'final_cleaned_data.csv' not found.")
            data = generate_dummy_data('final_cleaned_data.csv')
        else:
            data = pd.read_csv('final_cleaned_data.csv')
        
        X, y, encoders, proc_df = prepare_features_and_target(data)
        X_train, X_test, y_train, y_test, scaler = scale_and_split_data(X, y)
        results_df, best_cm, best_model_name, model_probs = train_and_evaluate_models(X_train, X_test, y_train, y_test)
        
        generate_plots(results_df, model_probs, y_test, best_cm, best_model_name)
        save_artifacts(X_train, X_test, y_train, y_test, results_df, best_cm, best_model_name)
    except Exception as e:
        print(f"An error occurred: {e}")
